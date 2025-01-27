from flask import Flask, request, jsonify, send_file, after_this_request
from threading import Thread
from werkzeug.utils import secure_filename
import cv2
from pyzbar.pyzbar import decode
import openpyxl
from datetime import datetime, date
import winsound
from flask_cors import CORS
import os
import threading
import time
import re
import requests
from bson import ObjectId  # Import for MongoDB ObjectId


from pymongo import MongoClient

# MongoDB connection
client = MongoClient("mongodb+srv://nehanshuparikh2006:pG1Ts0ZWZMSUKO1A@cluster0.xh9w5.mongodb.net/MajorProject?retryWrites=true&w=majority&appName=Cluster0")  # Replace with your MongoDB URI
db = client["MajorProject"]  # Database name
attendance_collection = db["attendances"]  # Collection name

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
CORS(app, expose_headers=["Content-Disposition"]) # This will enable CORS for all routes

def initialize_excel(level, school, branch, semester, division, subject, facultyName):
    # Consistent filename generation
    excel_file = f"{level}_{school}_{branch}_{semester}_{division}_{subject}.xlsx"

    try:
        workbook = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append([
            "User Type", "Enrollment", "Date", "Time",
            "Level", "Branch", "School", "Semester", "Division", "Subject", "Attendance Taken By"
        ])
        workbook.save(excel_file)

    return excel_file


def mark_attendance(barcode_data, name, level, school, branch, semester, division, subject, facultyName):
    # Generate consistent filename
    excel_file = initialize_excel(level, school, branch, semester, division, subject, facultyName)

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Attendance"]
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")

    # Avoid duplicate entries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] == barcode_data and row[2] == current_date:
            return False  # Duplicate entry

    # Extracting numeric part from semester and division
    semester_number = re.sub(r'\D', '', semester)  # Remove non-digit characters
    division_number = re.sub(r'\D', '', division)  # Remove non-digit characters

    
    # Add new entry
    sheet.append([
        name, barcode_data, current_date, current_time,
        level, branch, school, semester_number, division_number, subject, facultyName
    ])
    workbook.save(excel_file)
    return True

def convert_objectid_to_str(data):
    """Recursively converts ObjectId in a dictionary or list to string."""
    if isinstance(data, dict):
        return {key: convert_objectid_to_str(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [convert_objectid_to_str(item) for item in data]
    elif isinstance(data, ObjectId):
        return str(data)
    else:
        return data



def send_attendance_emails(attendance_data):
    # Send email to parents
    # Convert ObjectId to string in attendance data
    sanitized_data = convert_objectid_to_str(attendance_data)
    print(sanitized_data)
    # JavaScript backend URL
    backend_url = "http://localhost:5000/api/attendance/attendance-mail-to-parents"  # Replace with your backend URL
    try:
        # Send a POST request with attendance data
        response = requests.post(backend_url, json=sanitized_data)

        # Check response status
        if response.status_code == 200:
            print("Emails sent successfully!")
        else:
            print(f"Failed to send emails. Status Code: {response.status_code}, Response: {response.text}")

    except Exception as e:
        print(f"An error occurred while sending attendance emails: {str(e)}")



# Barcode scanning function that runs in a separate thread
def scan_barcode_in_thread(level, school, branch, semester, division, subject, facultyName):
    def scan_barcode():
        cap = cv2.VideoCapture(0)  # Open the webcam
        while True:
            ret, frame = cap.read()
            if not ret:
                continue
            for barcode in decode(frame):
                barcode_data = barcode.data.decode('utf-8')
                # Mark attendance
                if mark_attendance(
                    barcode_data, "Student", level, branch, school, semester, division, subject , facultyName
                ):
                    winsound.Beep(1000, 200)  # Success beep
                    print(f"Attendance marked for: {barcode_data}")
                     # Create attendance document
                    attendance_data = {
                        "userType": 'Student',
                        "enrollment": barcode_data,
                        "date": date.today().isoformat(),  # Store as ISO date string
                        "time": datetime.now().strftime("%H:%M:%S"),  # Current time
                        "level": level,
                        "branch": branch,
                        "school": school,
                        "semester": semester,
                        "division": division,
                        "subject": subject,
                        "attendanceTakenBy": facultyName
                    }
                    # Insert into MongoDB
                    attendance_collection.insert_one(attendance_data)
                else:
                    winsound.Beep(500, 200)  # Failure beep
                    print(f"Duplicate entry for: {barcode_data}")
            cv2.imshow("Barcode Scanner", frame)  # Display the video frame
            if cv2.waitKey(1) & 0xFF == ord('q'):  # Press 'q' to quit
                break
        cap.release()  # Release the webcam
        cv2.destroyAllWindows()  # Close the OpenCV window
        # Send email to parents after the scanner is turned off
        send_attendance_emails(attendance_data)

    # Start scanning in a separate thread
    Thread(target=scan_barcode, daemon=True).start()

# API endpoint to start attendance scanning
@app.route('/start-attendance', methods=['POST'])
def start_attendance():
    data = request.json
    level = data.get('level')
    school = data.get('school')
    branch = data.get('branch')
    semester = data.get('semester')
    division = data.get('division')
    subject = data.get('subject')
    facultyName = data.get('facultyName')

    # Validate that all required fields are provided
    if not all([level, school, branch, semester, division, subject, facultyName]):
        return jsonify({"message": "All fields are required!"}), 400

    # Start the barcode scanning in a thread
    scan_barcode_in_thread(level, school, branch, semester, division, subject, facultyName)

    return jsonify({"message": "Attendance started successfully!"}), 200


@app.route('/download-excel', methods=['GET'])
def download_excel():
    try:
        # Get query parameters
        level = request.args.get('level')
        school = request.args.get('school')
        branch = request.args.get('branch')
        semester = request.args.get('semester')
        division = request.args.get('division')
        subject = request.args.get('subject')
        
        print(f"[DEBUG] Received query parameters: level={level}, school={school}, branch={branch}, semester={semester}, division={division}, subject={subject}")

        # Validate that all parameters are present
        if not all([level, school, branch, semester, division, subject]):
            print("[DEBUG] Missing one or more required fields.")
            return jsonify({"message": "All fields are required."}), 400

        # Construct the file path
        file_name = f"{level}_{school}_{branch}_{semester}_{division}_{subject}.xlsx"
        file_path = os.path.join(BASE_DIR, file_name)
        print(f"[DEBUG] Full file path: {file_path}")

        # Check if the file exists
        if not os.path.exists(file_path):
            print(f"[DEBUG] File not found: {file_path}")
            return jsonify({"message": f"File '{file_name}' not found."}), 404

        # Validate the file format
        try:
            openpyxl.load_workbook(file_path)
            print(f"[DEBUG] File format validation successful for: {file_path}")
        except Exception as e:
            print(f"[DEBUG] Invalid Excel file format or corrupted file: {e}")
            return jsonify({"message": "File format is invalid or corrupted."}), 500

        # Send the file
        print(f"[DEBUG] Preparing to send file: {file_path}")

        # Delete the file in a background thread
        def remove_file():
            time.sleep(1)  # Wait for the file to be fully sent
            try:
                os.remove(file_path)
                print(f"[DEBUG] File {file_name} deleted successfully after download.")
            except Exception as e:
                print(f"[DEBUG] Error deleting file {file_name}: {e}")

        # Start the background task
        threading.Thread(target=remove_file).start()

        # Return the file for download
        return send_file(file_path, as_attachment=True, download_name=file_name)

    except Exception as e:
        # Log any unexpected errors
        print(f"[DEBUG] Unexpected error: {str(e)}")
        return jsonify({"message": f"Error downloading file: {str(e)}"}), 500

# Run the Flask app
if __name__ == '__main__':
    app.run(port=5001, threaded=True, debug=True)
